# -*- coding: utf-8 -*-

"""Fonctions utilitaires pour la lecture des fichiers XLS/XLSX ODK.

Ce module regroupe les helpers partagés entre l'algorithme principal
(formulaire_odk.py) : lecture des feuilles survey/choices, détection
des labels, types, groupes, etc.
"""

import os

from qgis.core import QgsProcessingException


# Préférence de langue : label::<fr> prioritaire, sinon 'label'
PREFERRED_LABEL_LANG = "fr"


# ----------------------------------------------------------------------
# Helpers XLS/XLSX
# ----------------------------------------------------------------------
def _choose_label(headers, lang=None):
    """Retourne le nom de la meilleure colonne de label."""
    low = {(h or "").lower(): (h or "") for h in headers}
    if lang:
        k = f"label::{lang}".lower()
        if k in low:
            return low[k]
    if "label" in low:
        return low["label"]
    for h in headers:
        if (h or "").lower().startswith("label::"):
            return h
    return None


def _read_xlsx(path, sheet_name="survey"):
    """Retourne (headers, rows) depuis la feuille sheet_name d'un .xlsx."""
    try:
        from openpyxl import load_workbook
    except Exception as e:
        # openpyxl manquant (rare en 3.28), on remonte une info claire
        raise QgsProcessingException(
            "Le module 'openpyxl' est requis pour lire les .xlsx. "
            "Installe-le dans l'environnement QGIS ou convertis le fichier en .xls."
        ) from e
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    headers = [c.value if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2):
        rows.append([c.value for c in row])
    return headers, rows


def _read_xls(path, sheet_name="survey"):
    """Retourne (headers, rows) depuis la feuille sheet_name d'un .xls."""
    try:
        import xlrd  # xlrd>=2.0 lit uniquement .xls (parfait ici)
    except ImportError:
        # xlrd absent : suggérer la conversion
        return [], []
    book = xlrd.open_workbook(path)
    try:
        sheet = book.sheet_by_name(sheet_name)
    except Exception:
        return [], []
    headers = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
    rows = []
    for r in range(1, sheet.nrows):
        rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    return headers, rows


def read_odk_mapping(path, preferred_lang=PREFERRED_LABEL_LANG):
    """Lit la feuille 'survey' et renvoie mapping {name -> label_proposé} + nom de la colonne de label utilisée."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        headers, rows = _read_xlsx(path, "survey")
    elif ext == ".xls":
        headers, rows = _read_xls(path, "survey")
    else:
        return {}, None

    headers = [str(h).strip() if h is not None else "" for h in headers]
    if not headers:
        return {}, None

    def find_col(col):
        for i, h in enumerate(headers):
            if h.lower() == col.lower():
                return i
        return None

    i_name = find_col("name")
    i_type = find_col("type")
    if i_name is None or i_type is None:
        return {}, None

    label_hdr = _choose_label(headers, preferred_lang)
    if not label_hdr:
        return {}, None
    i_label = headers.index(label_hdr)

    mapping = {}
    repeat_depth = 0
    for row in rows:
        get = lambda i: (row[i] if i is not None and i < len(row) else None)
        t = str(get(i_type) or "").strip().lower()

        if t in ("begin_repeat", "begin repeat"):
            repeat_depth += 1
            continue
        if t in ("end_repeat", "end repeat"):
            repeat_depth = max(0, repeat_depth - 1)
            continue
        if repeat_depth > 0:
            continue

        if t in ("begin_group", "begin group", "end_group", "end group"):
            continue

        nm = get(i_name)
        lb = get(i_label)
        if nm and (lb is not None):
            mapping[str(nm).strip()] = str(lb)

    return mapping, label_hdr


def read_odk_select_one(path, preferred_lang=PREFERRED_LABEL_LANG):
    """Retourne un dict {name: {"list_name": ..., "label": ...}} pour les questions select_one."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        headers, rows = _read_xlsx(path, "survey")
    elif ext == ".xls":
        headers, rows = _read_xls(path, "survey")
    else:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in headers]
    if not headers:
        return {}

    def find_col(col):
        for i, h in enumerate(headers):
            if h.lower() == col.lower():
                return i
        return None

    i_name = find_col("name")
    i_type = find_col("type")
    if i_name is None or i_type is None:
        return {}

    label_hdr = _choose_label(headers, preferred_lang)
    i_label = headers.index(label_hdr) if label_hdr else None

    out = {}
    repeat_depth = 0
    for row in rows:
        get = lambda i: (row[i] if i is not None and i < len(row) else None)
        raw_type = str(get(i_type) or "").strip()
        t = raw_type.lower()

        if t in ("begin_repeat", "begin repeat"):
            repeat_depth += 1
            continue
        if t in ("end_repeat", "end repeat"):
            repeat_depth = max(0, repeat_depth - 1)
            continue
        if repeat_depth > 0:
            continue

        if t in ("begin_group", "begin group", "end_group", "end group"):
            continue

        # On ne garde que les select_one <list_name>
        if not t.startswith("select_one"):
            continue

        parts = raw_type.split()
        if len(parts) < 2:
            continue
        list_name = parts[1].strip()
        if not list_name:
            continue

        nm = get(i_name)
        if not nm:
            continue

        lb = get(i_label) if i_label is not None else None
        out[str(nm).strip()] = {
            "list_name": list_name,
            "label": str(lb) if lb is not None else "",
        }

    return out


def read_odk_images(path):
    """Retourne l'ensemble des noms de champs dont le type est 'image' ou 'photo' dans la feuille survey."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        headers, rows = _read_xlsx(path, "survey")
    elif ext == ".xls":
        headers, rows = _read_xls(path, "survey")
    else:
        return set()

    headers = [str(h).strip() if h is not None else "" for h in headers]
    if not headers:
        return set()

    def find_col(col):
        for i, h in enumerate(headers):
            if h.lower() == col.lower():
                return i
        return None

    i_name = find_col("name")
    i_type = find_col("type")
    if i_name is None or i_type is None:
        return set()

    out = set()
    repeat_depth = 0
    for row in rows:
        get = lambda i: (row[i] if i is not None and i < len(row) else None)
        t = str(get(i_type) or "").strip().lower()

        if t in ("begin_repeat", "begin repeat"):
            repeat_depth += 1
            continue
        if t in ("end_repeat", "end repeat"):
            repeat_depth = max(0, repeat_depth - 1)
            continue
        if repeat_depth > 0:
            continue

        if t in ("begin_group", "begin group", "end_group", "end group"):
            continue

        # Accepte les types ODK "image" et "photo"
        if t not in ("image", "photo"):
            continue

        nm = get(i_name)
        if not nm:
            continue
        out.add(str(nm).strip())

    return out


def read_odk_basic_types(path):
    """Retourne un dict {name -> type de base ODK (text, integer, decimal, etc.)} pour la feuille survey."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        headers, rows = _read_xlsx(path, "survey")
    elif ext == ".xls":
        headers, rows = _read_xls(path, "survey")
    else:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in headers]
    if not headers:
        return {}

    def find_col(col):
        for i, h in enumerate(headers):
            if h.lower() == col.lower():
                return i
        return None

    i_name = find_col("name")
    i_type = find_col("type")
    if i_name is None or i_type is None:
        return {}

    types = {}
    repeat_depth = 0
    for row in rows:
        get = lambda i: (row[i] if i is not None and i < len(row) else None)
        raw_type = str(get(i_type) or "").strip()
        t = raw_type.lower()

        if t in ("begin_repeat", "begin repeat"):
            repeat_depth += 1
            continue
        if t in ("end_repeat", "end repeat"):
            repeat_depth = max(0, repeat_depth - 1)
            continue
        if repeat_depth > 0:
            continue

        if t in ("begin_group", "begin group", "end_group", "end group"):
            continue

        # type de base = premier mot (text, integer, decimal, select_one, image, photo, ...)
        base = t.split()[0] if t else ""
        nm = get(i_name)
        if not nm or not base:
            continue
        types[str(nm).strip()] = base

    return types


def read_odk_groups(path, preferred_lang=PREFERRED_LABEL_LANG):
    """Retourne un dict {field_name -> label_du_groupe} à partir de la feuille survey.

    On utilise les lignes begin_group / end_group pour construire une pile de groupes.
    Les repeats sont ignorés (autre table).
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        headers, rows = _read_xlsx(path, "survey")
    elif ext == ".xls":
        headers, rows = _read_xls(path, "survey")
    else:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in headers]
    if not headers:
        return {}

    def find_col(col):
        for i, h in enumerate(headers):
            if h.lower() == col.lower():
                return i
        return None

    i_name = find_col("name")
    i_type = find_col("type")
    if i_name is None or i_type is None:
        return {}

    label_hdr = _choose_label(headers, preferred_lang)
    i_label = headers.index(label_hdr) if label_hdr else None

    field_to_group = {}
    group_stack = []  # chaque entrée: {"name": ..., "label": ...}
    repeat_depth = 0  # on ignore entièrement les repeats pour les groupes

    for row in rows:
        get = lambda i: (row[i] if i is not None and i < len(row) else None)
        raw_type = str(get(i_type) or "").strip()
        t = raw_type.lower()

        # Gestion des repeats : on les exclut des groupes
        if t.startswith("begin_repeat") or t.startswith("begin repeat"):
            repeat_depth += 1
            continue
        if t.startswith("end_repeat") or t.startswith("end repeat"):
            repeat_depth = max(0, repeat_depth - 1)
            continue
        if repeat_depth > 0:
            continue

        # Début de groupe
        if t.startswith("begin_group") or t.startswith("begin group"):
            g_name = get(i_name)
            g_label = get(i_label) if i_label is not None else None
            label = str(g_label) if g_label not in (None, "") else (str(g_name) if g_name else "Groupe")
            group_stack.append({"name": str(g_name) if g_name else "", "label": label})
            continue

        # Fin de groupe
        if t.startswith("end_group") or t.startswith("end group"):
            if group_stack:
                group_stack.pop()
            continue

        nm = get(i_name)
        if not nm:
            continue

        if group_stack:
            current_group = group_stack[-1]
            field_to_group[str(nm).strip()] = current_group.get("label", "")

    return field_to_group


def read_odk_choices(path, preferred_lang=PREFERRED_LABEL_LANG):
    """Lit la feuille 'choices' et renvoie une liste de dicts
    [{"list_name": ..., "name": ..., "label": ...}, ...]."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        headers, rows = _read_xlsx(path, "choices")
    elif ext == ".xls":
        headers, rows = _read_xls(path, "choices")
    else:
        return []

    headers = [str(h).strip() if h is not None else "" for h in headers]
    if not headers:
        return []

    def find_col(col):
        for i, h in enumerate(headers):
            if h.lower() == col.lower():
                return i
        return None

    i_list_name = find_col("list_name")
    i_name = find_col("name")
    if i_list_name is None or i_name is None:
        return []

    label_hdr = _choose_label(headers, preferred_lang)
    i_label = headers.index(label_hdr) if label_hdr else None

    out = []
    for row in rows:
        get = lambda i: (row[i] if i is not None and i < len(row) else None)
        ln = get(i_list_name)
        nm = get(i_name)
        if not ln or not nm:
            continue
        lb = get(i_label) if i_label is not None else None
        out.append({
            "list_name": str(ln).strip(),
            "name": str(nm).strip(),
            "label": str(lb).strip() if lb not in (None, "") else "",
        })

    return out
